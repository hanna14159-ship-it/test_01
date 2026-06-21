import sqlite3
from copy import copy
from datetime import datetime
from io import BytesIO
from pathlib import Path
from zoneinfo import ZoneInfo

import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.units import pixels_to_EMU


# ==================================================
# 1. 기본 설정
# ==================================================

DB_PATH = "steamer.db"
TEMPLATE_PATH = "steamer_template.xlsx"
APPROVAL_IMAGE_PATH = "approval_box.png"

KST = ZoneInfo("Asia/Seoul")

MACHINES = [
    "11", "12", "13", "14",
    "21", "22", "23", "24",
    "51", "52", "53", "54",
]

# 11호기, 51호기만
# 증기 입구/중앙/출구 칸에
# 첫째 줄: 유량
# 둘째 줄: 압력
FLOW_PRESSURE_MACHINES = ["11", "51"]

# 현재 steamer_template.xlsx 기준 입력 행
MACHINE_ROW_MAP = {
    "11": 11,
    "12": 12,
    "13": 13,
    "14": 14,
    "21": 15,
    "22": 16,
    "23": 17,
    "24": 18,
    "51": 19,
    "52": 20,
    "53": 21,
    "54": 22,
}


# ==================================================
# 2. DB 함수
# ==================================================

def connect_db():
    return sqlite3.connect(DB_PATH)


def create_table():
    """
    점검 기록 테이블 생성.
    기존 DB에 새 칼럼이 없으면 자동 추가.
    """

    with connect_db() as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS steamer_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,

                check_date TEXT NOT NULL,
                check_time TEXT NOT NULL,
                machine TEXT NOT NULL,

                product TEXT,
                heat_temp REAL,
                oil_set_temp REAL,
                oil_now_temp REAL,
                steam_usage REAL,
                ct_water REAL,

                inlet_flow REAL,
                inlet_pressure REAL,

                middle_flow REAL,
                middle_pressure REAL,

                outlet_flow REAL,
                outlet_pressure REAL,

                memo TEXT,

                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,

                UNIQUE(check_date, machine)
            )
            """
        )

        existing_columns = [
            row[1]
            for row in conn.execute(
                "PRAGMA table_info(steamer_logs)"
            ).fetchall()
        ]

        new_columns = {
            "product": "TEXT",
            "heat_temp": "REAL",
            "oil_set_temp": "REAL",
            "oil_now_temp": "REAL",
            "steam_usage": "REAL",
            "ct_water": "REAL",
            "inlet_flow": "REAL",
            "inlet_pressure": "REAL",
            "middle_flow": "REAL",
            "middle_pressure": "REAL",
            "outlet_flow": "REAL",
            "outlet_pressure": "REAL",
            "memo": "TEXT",
            "created_at": "TEXT",
            "updated_at": "TEXT",
        }

        for column_name, column_type in new_columns.items():
            if column_name not in existing_columns:
                conn.execute(
                    f"ALTER TABLE steamer_logs "
                    f"ADD COLUMN {column_name} {column_type}"
                )

        now_text = datetime.now(KST).strftime("%Y-%m-%d %H:%M:%S")

        conn.execute(
            """
            UPDATE steamer_logs
            SET created_at = ?
            WHERE created_at IS NULL
               OR created_at = ''
            """,
            (now_text,),
        )

        conn.execute(
            """
            UPDATE steamer_logs
            SET updated_at = ?
            WHERE updated_at IS NULL
               OR updated_at = ''
            """,
            (now_text,),
        )


def save_or_update_record(record):
    """
    같은 날짜 + 같은 호기 기록이 없으면 새로 저장.
    이미 있으면 기존 기록 수정.
    """

    check_date = record[0]
    machine = record[2]

    with connect_db() as conn:
        existing_record = conn.execute(
            """
            SELECT id
            FROM steamer_logs
            WHERE check_date = ?
              AND machine = ?
            """,
            (check_date, machine),
        ).fetchone()

        if existing_record is None:
            conn.execute(
                """
                INSERT INTO steamer_logs (
                    check_date,
                    check_time,
                    machine,
                    product,
                    heat_temp,
                    oil_set_temp,
                    oil_now_temp,
                    steam_usage,
                    ct_water,
                    inlet_flow,
                    inlet_pressure,
                    middle_flow,
                    middle_pressure,
                    outlet_flow,
                    outlet_pressure,
                    memo,
                    created_at,
                    updated_at
                )
                VALUES (
                    ?, ?, ?, ?, ?, ?,
                    ?, ?, ?, ?, ?, ?,
                    ?, ?, ?, ?, ?, ?
                )
                """,
                record,
            )

            return "inserted"

        conn.execute(
            """
            UPDATE steamer_logs
            SET
                check_time = ?,
                product = ?,
                heat_temp = ?,
                oil_set_temp = ?,
                oil_now_temp = ?,
                steam_usage = ?,
                ct_water = ?,
                inlet_flow = ?,
                inlet_pressure = ?,
                middle_flow = ?,
                middle_pressure = ?,
                outlet_flow = ?,
                outlet_pressure = ?,
                memo = ?,
                updated_at = ?
            WHERE check_date = ?
              AND machine = ?
            """,
            (
                record[1],    # 점검 시간
                record[3],    # 제품명
                record[4],    # 열교환기 온도
                record[5],    # 유탕온도 설정
                record[6],    # 유탕온도 현재
                record[7],    # 증기 사용량
                record[8],    # C/T수
                record[9],    # 입구 유량
                record[10],   # 입구 압력
                record[11],   # 중앙 유량
                record[12],   # 중앙 압력
                record[13],   # 출구 유량
                record[14],   # 출구 압력
                record[15],   # 비고
                record[17],   # 수정 시각
                record[0],    # 날짜
                record[2],    # 호기
            ),
        )

        return "updated"


def load_records(selected_date):
    """
    선택한 날짜의 점검 기록을 불러온다.
    """

    date_text = selected_date.isoformat()

    conn = connect_db()
    conn.row_factory = sqlite3.Row

    records = conn.execute(
        """
        SELECT
            id,
            check_date,
            check_time,
            machine,
            product,
            heat_temp,
            oil_set_temp,
            oil_now_temp,
            steam_usage,
            ct_water,
            inlet_flow,
            inlet_pressure,
            middle_flow,
            middle_pressure,
            outlet_flow,
            outlet_pressure,
            memo,
            created_at,
            updated_at
        FROM steamer_logs
        WHERE check_date = ?
        ORDER BY CAST(machine AS INTEGER) ASC
        """,
        (date_text,),
    ).fetchall()

    conn.close()

    return records


def delete_record(record_id):
    """
    id 기준으로 점검 기록 한 건 삭제.
    """

    with connect_db() as conn:
        cursor = conn.execute(
            """
            DELETE FROM steamer_logs
            WHERE id = ?
            """,
            (record_id,),
        )

        return cursor.rowcount


# ==================================================
# 3. 입력값 처리 함수
# ==================================================

def parse_number(value, label):
    """
    문자 입력값을 숫자로 변환한다.

    빈칸은 None으로 저장.
    소수점 아래 5자리까지 허용.
    """

    value = value.strip()

    if value == "":
        return None

    try:
        number = float(value)
    except ValueError:
        raise ValueError(f"{label}은 숫자만 입력해야 합니다.")

    if "." in value:
        decimal_part = value.split(".")[1]

        if len(decimal_part) > 5:
            raise ValueError(
                f"{label}은 소수점 아래 5자리까지만 입력 가능합니다."
            )

    return number


def format_number(value):
    """
    Excel 한 칸 안에 들어갈 숫자를 보기 좋게 변환.
    None이면 빈칸.
    """

    if value is None:
        return ""

    if isinstance(value, float):
        return f"{value:.5f}".rstrip("0").rstrip(".")

    return str(value)


def number_text_input(label, key):
    """
    + / - 버튼 없는 숫자 입력칸.
    """

    return st.text_input(
        label,
        key=key,
        placeholder="예: 123.45",
    )


# ==================================================
# 4. QR 주소에서 호기 읽기
# ==================================================

def get_machine_from_url():
    """
    URL의 machine 값을 읽는다.

    예:
    ?machine=11
    """

    machine = st.query_params.get("machine")

    if machine in MACHINES:
        return machine

    return None


# ==================================================
# 5. Excel 보조 함수
# ==================================================

def get_writable_cell_address(worksheet, cell_address):
    """
    병합셀인 경우 실제로 값을 쓸 수 있는 왼쪽 위 셀 주소를 반환한다.
    병합셀이 아니면 원래 셀 주소를 그대로 반환한다.
    """

    for merged_range in worksheet.merged_cells.ranges:
        if cell_address in merged_range:
            return merged_range.start_cell.coordinate

    return cell_address


def write_cell_value(worksheet, cell_address, value):
    """
    병합셀 오류를 방지하면서 값을 입력한다.
    병합된 셀 안쪽이면 병합 범위의 왼쪽 위 셀에 값을 넣는다.
    """

    writable_cell_address = get_writable_cell_address(
        worksheet,
        cell_address,
    )

    worksheet[writable_cell_address] = value


def make_flow_pressure_text(flow, pressure):
    """
    유량과 압력을 한 셀에 두 줄로 넣는다.

    첫째 줄: 유량
    둘째 줄: 압력
    """

    flow_text = format_number(flow)
    pressure_text = format_number(pressure)

    if flow_text == "" and pressure_text == "":
        return ""

    return f"{flow_text}\n{pressure_text}"


def write_multiline_cell(worksheet, cell_address, value):
    """
    병합셀 오류를 방지하면서 두 줄 값을 입력한다.
    원본 테두리, 행높이, 열너비는 건드리지 않는다.
    """

    writable_cell_address = get_writable_cell_address(
        worksheet,
        cell_address,
    )

    cell = worksheet[writable_cell_address]
    cell.value = value

    old_alignment = copy(cell.alignment)

    cell.alignment = Alignment(
        horizontal=old_alignment.horizontal,
        vertical=old_alignment.vertical,
        text_rotation=old_alignment.text_rotation,
        wrap_text=True,
        shrink_to_fit=old_alignment.shrink_to_fit,
        indent=old_alignment.indent,
    )


def column_width_to_pixels(width):
    """
    Excel 열 너비를 대략적인 픽셀 값으로 변환.
    """

    if width is None:
        width = 8.43

    return int(width * 7 + 5)


def find_image_start_from_top_right(worksheet, top_right_cell, image_width_px):
    """
    top_right_cell의 오른쪽 위 모서리를 기준으로
    이미지의 오른쪽 위 모서리를 맞추기 위한 시작 위치 계산.
    """

    col_text = "".join([c for c in top_right_cell if c.isalpha()])
    row_text = "".join([c for c in top_right_cell if c.isdigit()])

    right_col_idx = column_index_from_string(col_text)
    row_idx = int(row_text)

    remaining_width = image_width_px
    current_col_idx = right_col_idx

    while current_col_idx >= 1:
        col_letter = get_column_letter(current_col_idx)
        col_width_px = column_width_to_pixels(
            worksheet.column_dimensions[col_letter].width
        )

        if remaining_width <= col_width_px:
            start_offset_px = col_width_px - remaining_width
            return current_col_idx, start_offset_px, row_idx

        remaining_width -= col_width_px
        current_col_idx -= 1

    return 1, 0, row_idx


def insert_approval_image_top_right(
    worksheet,
    image_path,
    top_right_cell="K5",
    image_width_px=220,
    image_height_px=60,
):
    """
    approval_box.png를 삽입하되,
    이미지의 오른쪽 위 모서리를
    top_right_cell의 오른쪽 위 모서리에 맞춘다.
    """

    if not Path(image_path).exists():
        return

    start_col_idx, start_offset_px, row_idx = find_image_start_from_top_right(
        worksheet,
        top_right_cell,
        image_width_px,
    )

    img = Image(image_path)
    img.width = image_width_px
    img.height = image_height_px

    marker = AnchorMarker(
        col=start_col_idx - 1,
        colOff=pixels_to_EMU(start_offset_px),
        row=row_idx - 1,
        rowOff=0,
    )

    size = XDRPositiveSize2D(
        cx=pixels_to_EMU(image_width_px),
        cy=pixels_to_EMU(image_height_px),
    )

    img.anchor = OneCellAnchor(
        _from=marker,
        ext=size,
    )

    worksheet.add_image(img)


# ==================================================
# 6. Excel 생성 함수
# ==================================================

def make_template_excel(selected_date):
    """
    steamer_template.xlsx 원본 양식을 열어서
    값만 지정된 셀에 입력한다.

    결재칸은 approval_box.png를 K5 우측 상단 모서리 기준으로 삽입한다.
    """

    if not Path(TEMPLATE_PATH).exists():
        raise FileNotFoundError(
            f"{TEMPLATE_PATH} 파일을 찾을 수 없습니다. "
            "GitHub 저장소에 steamer_template.xlsx 파일을 올렸는지 확인하세요."
        )

    records = load_records(selected_date)

    workbook = load_workbook(TEMPLATE_PATH)
    worksheet = workbook.active

    weekday_names = [
        "월", "화", "수", "목", "금", "토", "일"
    ]

    weekday = weekday_names[selected_date.weekday()]

    # 날짜 입력 위치
    write_cell_value(
        worksheet,
        "A6",
        (
            f"  {selected_date.year} 년       "
            f"{selected_date.month} 월       "
            f"{selected_date.day} 일       "
            f"{weekday}요일"
        ),
    )

    # 결재칸 그림 삽입
    insert_approval_image_top_right(
        worksheet,
        APPROVAL_IMAGE_PATH,
        top_right_cell="K5",
        image_width_px=220,
        image_height_px=60,
    )

    for record in records:
        machine = record["machine"]

        if machine not in MACHINE_ROW_MAP:
            continue

        row = MACHINE_ROW_MAP[machine]

        write_cell_value(worksheet, f"B{row}", record["check_time"])
        write_cell_value(worksheet, f"C{row}", record["product"])
        write_cell_value(worksheet, f"D{row}", record["heat_temp"])
        write_cell_value(worksheet, f"E{row}", record["oil_set_temp"])
        write_cell_value(worksheet, f"F{row}", record["oil_now_temp"])

        # 증기 사용량은 값이 없으면 엑셀에 "." 입력
        if record["steam_usage"] is None:
            write_cell_value(worksheet, f"G{row}", ".")
        else:
            write_cell_value(worksheet, f"G{row}", record["steam_usage"])

        write_cell_value(worksheet, f"H{row}", record["ct_water"])

        if machine in FLOW_PRESSURE_MACHINES:
            write_multiline_cell(
                worksheet,
                f"K{row}",
                make_flow_pressure_text(
                    record["inlet_flow"],
                    record["inlet_pressure"],
                ),
            )

            write_multiline_cell(
                worksheet,
                f"M{row}",
                make_flow_pressure_text(
                    record["middle_flow"],
                    record["middle_pressure"],
                ),
            )

            write_multiline_cell(
                worksheet,
                f"O{row}",
                make_flow_pressure_text(
                    record["outlet_flow"],
                    record["outlet_pressure"],
                ),
            )

        else:
            write_cell_value(
                worksheet,
                f"K{row}",
                record["inlet_pressure"],
            )

            write_cell_value(
                worksheet,
                f"M{row}",
                record["middle_pressure"],
            )

            write_cell_value(
                worksheet,
                f"O{row}",
                record["outlet_pressure"],
            )

    # 비고는 현재 Excel에는 입력하지 않음.
    # 앱 저장 기록에는 비고가 그대로 저장됨.

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return output.getvalue(), len(records)


# ==================================================
# 7. 앱 초기 설정
# ==================================================

st.set_page_config(
    page_title="증숙기 점검일지",
    page_icon="📝",
    layout="wide",
)

create_table()

qr_machine = get_machine_from_url()

if qr_machine is not None:
    default_machine_index = MACHINES.index(qr_machine)
else:
    default_machine_index = 0


# ==================================================
# 8. 화면 구성
# ==================================================

st.title("증숙기 점검일지")

tab_input, tab_records, tab_manage, tab_excel = st.tabs(
    [
        "점검 입력",
        "저장 기록",
        "기록 관리",
        "Excel 다운로드",
    ]
)


# ==================================================
# 9. 점검 입력 탭
# ==================================================

with tab_input:

    st.subheader("증숙기 점검값 입력")

    now = datetime.now(KST)

    st.info(
        f"점검 날짜/시간은 저장 버튼을 누르는 현재 시각으로 자동 저장됩니다. "
        f"현재 기준: {now.strftime('%Y-%m-%d %H:%M')}"
    )

    if qr_machine is not None:
        st.success(
            f"QR을 통해 {qr_machine}호기가 자동 선택되었습니다."
        )
    else:
        st.info(
            "QR 접속이 아닙니다. 호기를 직접 선택해 주세요."
        )

    machine = st.selectbox(
        "호기 선택",
        MACHINES,
        index=default_machine_index,
    )

    product = st.text_input("제품명")

    heat_temp_text = number_text_input(
        "열교환기 온도",
        "heat_temp",
    )

    oil_set_temp_text = number_text_input(
        "유탕온도 설정",
        "oil_set_temp",
    )

    oil_now_temp_text = number_text_input(
        "유탕온도 현재",
        "oil_now_temp",
    )

    steam_usage_text = number_text_input(
        "증기 사용량",
        "steam_usage",
    )

    ct_water_text = number_text_input(
        "C/T수",
        "ct_water",
    )

    if machine in FLOW_PRESSURE_MACHINES:
        st.markdown("#### 증기 입구 / 중앙 / 출구")

        inlet_flow_text = number_text_input(
            "입구 유량",
            "inlet_flow",
        )

        inlet_pressure_text = number_text_input(
            "입구 압력",
            "inlet_pressure",
        )

        middle_flow_text = number_text_input(
            "중앙 유량",
            "middle_flow",
        )

        middle_pressure_text = number_text_input(
            "중앙 압력",
            "middle_pressure",
        )

        outlet_flow_text = number_text_input(
            "출구 유량",
            "outlet_flow",
        )

        outlet_pressure_text = number_text_input(
            "출구 압력",
            "outlet_pressure",
        )

    else:
        inlet_flow_text = ""
        middle_flow_text = ""
        outlet_flow_text = ""

        inlet_pressure_text = number_text_input(
            "입구 압력",
            "inlet_pressure",
        )

        middle_pressure_text = number_text_input(
            "중앙 압력",
            "middle_pressure",
        )

        outlet_pressure_text = number_text_input(
            "출구 압력",
            "outlet_pressure",
        )

    memo = st.text_area("비고")

    submitted = st.button(
        "저장",
        use_container_width=True,
    )

    if submitted:

        try:
            save_time = datetime.now(KST)

            check_date = save_time.date()
            check_time = save_time.strftime("%H:%M")
            current_time = save_time.strftime("%Y-%m-%d %H:%M:%S")

            heat_temp = parse_number(
                heat_temp_text,
                "열교환기 온도",
            )

            oil_set_temp = parse_number(
                oil_set_temp_text,
                "유탕온도 설정",
            )

            oil_now_temp = parse_number(
                oil_now_temp_text,
                "유탕온도 현재",
            )

            steam_usage = parse_number(
                steam_usage_text,
                "증기 사용량",
            )

            ct_water = parse_number(
                ct_water_text,
                "C/T수",
            )

            inlet_flow = parse_number(
                inlet_flow_text,
                "입구 유량",
            )

            inlet_pressure = parse_number(
                inlet_pressure_text,
                "입구 압력",
            )

            middle_flow = parse_number(
                middle_flow_text,
                "중앙 유량",
            )

            middle_pressure = parse_number(
                middle_pressure_text,
                "중앙 압력",
            )

            outlet_flow = parse_number(
                outlet_flow_text,
                "출구 유량",
            )

            outlet_pressure = parse_number(
                outlet_pressure_text,
                "출구 압력",
            )

            record = (
                check_date.isoformat(),
                check_time,
                machine,
                product.strip(),
                heat_temp,
                oil_set_temp,
                oil_now_temp,
                steam_usage,
                ct_water,
                inlet_flow,
                inlet_pressure,
                middle_flow,
                middle_pressure,
                outlet_flow,
                outlet_pressure,
                memo.strip(),
                current_time,
                current_time,
            )

            result = save_or_update_record(record)

            if result == "inserted":
                st.success(
                    f"{machine}호기 점검 기록이 새로 저장되었습니다. "
                    f"저장시간: {current_time}"
                )

            elif result == "updated":
                st.success(
                    f"{machine}호기 기존 기록이 수정되었습니다. "
                    f"수정시간: {current_time}"
                )

        except ValueError as error:
            st.error(str(error))

        except sqlite3.Error as error:
            st.error(
                f"저장 또는 수정 중 오류가 발생했습니다: {error}"
            )


# ==================================================
# 10. 저장 기록 탭
# ==================================================

with tab_records:

    st.subheader("날짜별 저장 기록")

    search_date = st.date_input(
        "조회 날짜",
        key="search_date",
    )

    records = load_records(search_date)

    if not records:
        st.info(
            f"{search_date.isoformat()}에 저장된 기록이 없습니다."
        )

    else:
        display_records = []

        for record in records:
            display_records.append(
                {
                    "번호": record["id"],
                    "날짜": record["check_date"],
                    "시간": record["check_time"],
                    "호기": record["machine"],
                    "제품명": record["product"],
                    "열교환기온도": record["heat_temp"],
                    "유탕온도 설정": record["oil_set_temp"],
                    "유탕온도 현재": record["oil_now_temp"],
                    "증기사용량": record["steam_usage"],
                    "C/T수": record["ct_water"],
                    "입구유량": record["inlet_flow"],
                    "입구압력": record["inlet_pressure"],
                    "중앙유량": record["middle_flow"],
                    "중앙압력": record["middle_pressure"],
                    "출구유량": record["outlet_flow"],
                    "출구압력": record["outlet_pressure"],
                    "비고": record["memo"],
                    "최초 등록": record["created_at"],
                    "마지막 수정": record["updated_at"],
                }
            )

        st.write(f"총 {len(display_records)}건")

        st.dataframe(
            display_records,
            use_container_width=True,
            hide_index=True,
        )


# ==================================================
# 11. 기록 관리 탭
# ==================================================

with tab_manage:

    st.subheader("저장 기록 삭제")

    manage_date = st.date_input(
        "관리할 날짜",
        key="manage_date",
    )

    manage_records = load_records(manage_date)

    if not manage_records:
        st.info(
            f"{manage_date.isoformat()}에 저장된 기록이 없습니다."
        )

    else:
        record_options = {}

        for record in manage_records:
            label = (
                f'{record["id"]}번 | '
                f'{record["machine"]}호기 | '
                f'{record["check_time"]} | '
                f'{record["product"]}'
            )

            record_options[label] = record["id"]

        selected_label = st.selectbox(
            "삭제할 기록",
            list(record_options.keys()),
        )

        selected_record_id = record_options[selected_label]

        delete_confirmed = st.checkbox(
            "선택한 기록을 삭제하는 것에 동의합니다."
        )

        if st.button(
            "선택 기록 삭제",
            type="primary",
            use_container_width=True,
        ):

            if not delete_confirmed:
                st.warning(
                    "삭제 확인 체크박스를 선택해 주세요."
                )

            else:
                try:
                    deleted_count = delete_record(
                        selected_record_id
                    )

                    if deleted_count == 1:
                        st.success(
                            "선택한 기록이 삭제되었습니다."
                        )

                        st.rerun()

                    else:
                        st.warning(
                            "삭제할 기록을 찾을 수 없습니다."
                        )

                except sqlite3.Error as error:
                    st.error(
                        f"삭제 중 오류가 발생했습니다: {error}"
                    )


# ==================================================
# 12. Excel 다운로드 탭
# ==================================================

with tab_excel:

    st.subheader("기존 증숙기 점검일지 생성")

    excel_date = st.date_input(
        "점검일지 생성 날짜",
        key="excel_date",
    )

    records = load_records(excel_date)

    if not records:
        st.warning(
            f"{excel_date.isoformat()}에 저장된 기록이 없습니다."
        )

    else:
        try:
            excel_file, machine_count = make_template_excel(
                excel_date
            )

            file_name = (
                f"증숙기_점검일지_"
                f"{excel_date.isoformat()}.xlsx"
            )

            st.write(
                f"{machine_count}개 호기의 기록이 "
                "Excel 양식에 입력됩니다."
            )

            st.download_button(
                label="기존 양식 Excel 다운로드",
                data=excel_file,
                file_name=file_name,
                mime=(
                    "application/vnd.openxmlformats-"
                    "officedocument.spreadsheetml.sheet"
                ),
                use_container_width=True,
            )

        except FileNotFoundError as error:
            st.error(str(error))

        except Exception as error:
            st.error(
                f"Excel 생성 중 오류가 발생했습니다: {error}"
            )
